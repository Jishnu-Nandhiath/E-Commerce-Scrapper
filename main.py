import urllib.request
from urllib.error import HTTPError
from bs4 import BeautifulSoup
import pandas as pd
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
import time
from collections import OrderedDict

USER_AGENT = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.1916.47 Safari/537.36'

""" 
Required fields, Category, Name, Variant Name, Price, Image URL, Description
"""

"""  
from pixie website, I can take category, Name, Price and Image URL, as there is no Description for the product there
"""

def get_page(url):
    full_url = url.replace(" ", "%20")
    # fetch data points
    req = urllib.request.Request(
        full_url,
        data=None,
        headers={
            'User-Agent': USER_AGENT
        }
    )
    while True:
        try:
            response = urllib.request.urlopen(req)
            response_code = response.getcode()
            data= response.read()
            break
        except Exception as e:
            print(str(e))
            print('Blocked! Sleeping...')
            print('Retrying')
    html_page = data.decode('utf-8', errors='replace')
    return html_page


def productYield(mainUrl):
    html_page = get_page(mainUrl)
    if(not html_page):
        return
    
    soup = BeautifulSoup(html_page, 'html.parser')
    product_list_items = soup.find_all('li', class_='item product product-item')
    
    for product in product_list_items:
        output_dict = OrderedDict()
        
        # Extract product name
        product_name_element = product.find('a', class_='product-item-link')
        product_name = product_name_element.text.strip() if product_name_element else None
        output_dict['Product Name'] = product_name
        
        # Extract Product Link
        product_link = product_name_element['href'] if product_name_element else None
        output_dict['Product URL'] = product_link
        
        # Find the old price value
        old_price_span = product.find('span', class_='old-price')
        old_price = old_price_span.find('span', class_='price-wrapper')['data-price-amount'] if old_price_span else None
        output_dict['MRP'] = float(old_price) if old_price else None
        
        output_dict['Offer Price'] = None
        
        if(old_price_span is None):
            max_price_element = product.find('span', class_='price-container')
            max_price = max_price_element.find('span', class_='price-wrapper')['data-price-amount'] if max_price_element else None
            output_dict['MRP'] = float(max_price)
        elif(old_price_span is not None):
            # Find the special price value
            special_price_span = product.find('span', class_='special-price')
            special_price = special_price_span.find('span', class_='price-wrapper')['data-price-amount'] if special_price_span else None    
            output_dict['Offer Price']  = float(special_price)
        
        yield output_dict
        
    # return product main details, 
 
def productImageScrapper(product_url):
    driver = webdriver.Chrome()
    
    # this is cache key for image with 1200x1200 spec
    pixie_high_res_image_cache_key = "be23c90f8928c29b64f1808a247a0a32"
    
    image_url_list = list()
    
    max_retries = 3
    retry_count = 0

    while retry_count < max_retries:
        try:
            driver.get(product_url)

            wait = WebDriverWait(driver, 20)
            image_nav_shaft_div = wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'fotorama__nav__shaft')))
            
            fotorama_content = image_nav_shaft_div.get_attribute('innerHTML')
            soup = BeautifulSoup(fotorama_content, 'html.parser')
            
            # if img tags are found, then there is more than one image in the website, else only one image is available for the product
            if(soup.find('img')):
                img_tags = soup.find_all('img')
                if img_tags:
                    for img in img_tags:
                        img_src = img["src"]
                        img_src = img_src.replace('/cache/550ac8ae667c3d62b135794f574936e3/', f'/cache/{pixie_high_res_image_cache_key}/')
                        image_url_list.append(img_src)
                    print(image_url_list)
                    break
            else:
                soup = BeautifulSoup(driver.page_source, 'html.parser')
                main_image = soup.find('div', class_ = 'fotorama__stage__frame')
                main_image_url = main_image.find('img')['src']
                main_image_url = main_image_url.replace('/cache/39500cf9d88472c26e1ec72addc15f3a/', f'/cache/{pixie_high_res_image_cache_key}/')
                image_url_list.append(main_image_url)
                break
            
        except TimeoutException:
            # Handle a timeout exception (element not found)
            print("Element not found. Retrying...")
        print(f"Image try count :{retry_count}")
        retry_count += 1
        time.sleep(2)  # Wait for a short time before retrying

    if retry_count == max_retries:
        print("Max retries reached. Element not found.")

    # Close the WebDriver when done
    driver.quit()
    
    return image_url_list


def productScrapper(url):
    html_page = get_page(url)
    if(not html_page):
        return
    
    soup = BeautifulSoup(html_page, 'html.parser')
    
    output_dict = OrderedDict()
    
    #  ************************* Product Main Description **************************

    main_description_div = soup.find('div', class_='product attribute overview')
    
    output_dict["Main Description"] = None
    
    if main_description_div:
        main_description_ul = main_description_div.find('ul')
        if main_description_ul:
            main_description_li_tags = main_description_ul.find_all('li')
            if main_description_li_tags:
                main_description_text = "\n".join([li.get_text(strip=True) for li in main_description_li_tags])

                output_dict["Main Description"] = main_description_text
    
    
    """ 
    **********************************  Fetch Images **********************************
    
    image url sample : https://pixiesmediapull-145ca.kxcdn.com/pub/media/catalog/product/cache/39500cf9d88472c26e1ec72addc15f3a/a/r/aroma-magic-3-in-1-jasmine-blossom_5.jpeg
    if there is multiple images in the product page, then the cache key is same for the same image dimensions, only the url image name endpoint need to be fetched
    but this segment is dynamically loaded, hence may need to use selenium
    """
    output_dict["Image URLs"] = None
    image_list = productImageScrapper(url)
    output_dict["Image URLs"] = ",".join(image_list)
    
    # **********************************  Details  *********************************
    output_dict["Details"] = None
    
    product_details_div = soup.find('div', class_= 'product attribute description')
    formatted_details_text = ''
    if product_details_div:
        # print(product_details_div.prettify())
        # for br_tag in product_details_div.find_all('br'):
        #     br_tag.extract()
        # print(product_details_div.prettify())
        # for br_tag in product_details_div.find_all(['br', '/br', 'br/']):
        #     br_tag.unwrap()
        formatted_details_list = list()
        for element in product_details_div.find_all(['p', 'h3', 'li']):
            text = element.get_text(strip=True)
            if text not in formatted_details_list:
                formatted_details_list.append(text)
                
                if element.name == 'h3':
                    formatted_details_text += "\n\n" + element.get_text(strip=True) + ' '
                else:
                    for strong in element.find_all('strong'):
                        formatted_details_text += '\n\n' + strong.get_text(strip=True) + ' '
                        strong.extract()
                    formatted_details_text += '\n' + element.get_text(strip=True) + ' ' # Separate h3 tags with a new line


    output_dict["Details"] = formatted_details_text
    
    # **********************************  More Information  *********************************
    
    more_information_table = soup.find('table','data table additional-attributes')
    more_information_table_data_pairs = list()
    
    output_dict['More Information'] = None
    
    if more_information_table:
        for table_row in more_information_table.find_all('tr'):
            header = table_row.find('th', class_='col label').text.strip()
            data = table_row.find('td', class_='col data').text.strip()
            more_information_table_data_pairs.append(f"{header} : {data}")
    
        # join the table
        more_information_text = '\n'.join(more_information_table_data_pairs)

        output_dict['More Information'] = more_information_text
        
    return output_dict

# import requests  
# a = requests.get("https://pixies.in/aroma-magic-jasmine-body-wash")

# with open('a.txt','w', encoding= 'utf-8') as f:
#     f.write(a.text)

# mainScrapper("https://pixies.in/aroma-magic-jasmine-body-wash")

# def write_products(file, )

def fetchProductNumber(htmlcode):
    soup = BeautifulSoup(htmlcode, 'html.parser')

    # Find the element that contains the current page number
    element_count = 0
    current_page_element = soup.find('li', class_='item current')
    last_page_li_elements = soup.findAll('li', class_='item product product-item')
    element_count = len(last_page_li_elements)
    
    total_product_count = element_count
    # Extract the page number
    if current_page_element:
        current_page_text = current_page_element.text.strip()
        page_number = current_page_text.split()[-1]
        total_product_count = (int(page_number) - 1) * 30
        total_product_count += element_count

        print("Current Page Number:", page_number, "\n Total Product Count:", total_product_count)
    else:
        print("Current page element not found.")
    return total_product_count      



def write_excel(row_input):
    write_excel_file_path = 'output.xlsx'
    sheet_name = 'Sheet1'
    
    existing_data = pd.read_excel(write_excel_file_path, sheet_name=sheet_name)
    print(existing_data)
    
    # print(row_input)
        
    # print(row_input.keys())
    
    # print(existing_data.columns)
        
    new_data = pd.DataFrame([row_input])
    
    # combined_data = pd.concat([existing_data, df], ignore_index=True)
    print(new_data)
    
    df = pd.concat([existing_data, new_data])
    
    print(df)
    
    # existing_data.loc[len(existing_data)] = new_data
    
    df.to_excel(write_excel_file_path, sheet_name=sheet_name, index=False, header= False)
    
    
    
    
    
from openpyxl import load_workbook

# order : Product Name, Product URL, MRP, Offer Price, Brand, Main Description, Image URLs, Details, More Information
    
def append_excel(row_input):
    excel_file_path = 'output.xlsx'

    # Replace 'Sheet1' with the name of the sheet containing your data
    sheet_name = 'Sheet1'
    
    # book = load_workbook(excel_file_path)
    writer = pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') 
    writer.workbook = load_workbook(excel_file_path)
    
    df = pd.DataFrame([row_input])
    
    print("max row : ",writer.sheets['Sheet1'].max_row)
    
    df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)

    writer.close()
    
    
def scrapController():
    excel_file_path = 'pixies.xlsx'

    # Replace 'Sheet1' with the name of the sheet containing your data
    sheet_name = 'Sheet1'

    # Load the Excel file into a pandas DataFrame
    df = pd.read_excel(excel_file_path, sheet_name=sheet_name)
    
    output_index = 1
    
    for index, row in df.iterrows():
        brand_name = row['Brand Name']
        brand_link = row['Link']
        product_count = row['No of Products']
        
        if(index < 10):
            continue
        
        # product count is passed as 120, so to get page count -> it is divided by 120 
        page_count = int(product_count / 120) + 1
        
        # product main url sample : https://pixies.in/brands/aroma-magic?p=2&product_list_limit=120
        
        for page in range(1, page_count + 1):
            page_main_url = brand_link + f"?p={page}&product_list_limit=120"
            for mainProductDetails in productYield(page_main_url):
                print("output_index :" , output_index)
                currentProductRow = mainProductDetails
                currentProductRow["Brand"] = brand_name if brand_name else None
                productInformationDict = productScrapper(mainProductDetails["Product URL"])
                currentProductRow.update(productInformationDict)
                append_excel(currentProductRow)
                output_index += 1
                

                
scrapController()  

# productScrapper("https://pixies.in/brazilian-hairtech-capacabana-brazilian-protein-tanino-botox-1000ml")              


    
    
    # use get else dictionary key error can happen
    